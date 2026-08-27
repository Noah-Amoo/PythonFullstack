import io
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Any

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=11)
LABEL_FONT = Font(name="Calibri", size=11, bold=True)


def _sanitize_value(value: Any) -> Any:
    """
    Makes a raw DB value safe for openpyxl to write into a cell.

    - Excel/openpyxl has no concept of timezones: a tz-aware
      datetime (e.g. from a Postgres TIMESTAMPTZ column) raises
      TypeError on write. Convert to UTC first, then drop tzinfo,
      so the wall-clock value written is still correct UTC time
      rather than being silently reinterpreted as local time.
    - Decimal (common for money columns) isn't natively written by
      openpyxl in older versions and prints ugly in some readers —
      convert to float.
    """

    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)
        return value

    if isinstance(value, date):
        return value

    if isinstance(value, Decimal):
        return float(value)

    return value


def _autofit_columns(sheet, rows: list[dict[str, Any]]) -> None:
    """Roughly sizes each column to its widest value, capped so one
    long cell (e.g. a UUID or note) doesn't blow out the sheet."""

    if not rows:
        return

    headers = list(rows[0].keys())

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            value = row.get(header, "")
            max_len = max(max_len, len(str(value)))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)


def build_query_result_workbook(
    rows: list[dict[str, Any]],
    question: str = "",
    sql_query: str = "",
    answer: str = "",
) -> io.BytesIO:
    """
    Builds an in-memory .xlsx workbook from the SQL agent's query
    result.

    Sheet 1 ("Result"): one row per record, one column per field —
    whatever columns the SQL SELECT returned.
    Sheet 2 ("Details"): the original admin question, the SQL the
    agent generated, and its natural-language answer, so the export
    is traceable back to what was asked.
    """

    workbook = Workbook()

    # ---------------- Result sheet ----------------
    result_sheet = workbook.active
    result_sheet.title = "Result"

    if rows:
        headers = list(rows[0].keys())

        for col_idx, header in enumerate(headers, start=1):
            cell = result_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                cell = result_sheet.cell(
                    row=row_idx,
                    column=col_idx,
                    value=_sanitize_value(row.get(header)),
                )
                cell.font = BODY_FONT

        result_sheet.freeze_panes = "A2"
        _autofit_columns(result_sheet, rows)
    else:
        empty_cell = result_sheet.cell(row=1, column=1, value="No rows returned")
        empty_cell.font = BODY_FONT

    # ---------------- Details sheet ----------------
    details_sheet = workbook.create_sheet("Details")

    details = [
        ("Question", question),
        ("Generated SQL", sql_query),
        ("Answer", answer),
        (
            "Generated at (UTC)",
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
        ),
    ]

    for row_idx, (label, value) in enumerate(details, start=1):
        details_sheet.cell(row=row_idx, column=1, value=label).font = LABEL_FONT
        details_sheet.cell(row=row_idx, column=2, value=value).font = BODY_FONT

    details_sheet.column_dimensions["A"].width = 20
    details_sheet.column_dimensions["B"].width = 100

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return buffer


def excel_streaming_response(
    rows: list[dict[str, Any]],
    question: str = "",
    sql_query: str = "",
    answer: str = "",
    filename_prefix: str = "admin_query_result",
) -> StreamingResponse:
    """Wraps build_query_result_workbook as a downloadable FastAPI response."""

    buffer = build_query_result_workbook(
        rows=rows,
        question=question,
        sql_query=sql_query,
        answer=answer,
    )

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"{filename_prefix}_{timestamp}.xlsx"

    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )